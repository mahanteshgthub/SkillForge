import os
from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

def generate_workorder_pdf(workorder, template_path, output_pdf_path):
    # Load the Excel template
    wb = load_workbook(template_path)
    ws = wb.active

    # Extract values based on your provided cell mapping
    trade = ws["C3"].value or ""
    year = ws["E3"].value or ""
    ex_no = ws["Q3"].value or ""
    aim = ws["U3"].value or ""
    tolerance = ws["Z16"].value or ""

    # Feature extraction from Excel ranges
    def extract_range_values(col_letter, row_start, row_end):
        return [ws[f"{col_letter}{i}"].value or "" for i in range(row_start, row_end + 1)]

    size_list = extract_range_values("S", 6, 15)
    size_marks = extract_range_values("AD", 6, 15)
    subj_ops = extract_range_values("S", 20, 24)
    subj_marks = extract_range_values("AB", 20, 24)
    attitude_traits = extract_range_values("S", 27, 30)
    attitude_marks = extract_range_values("Z", 27, 30)

    # Prepare PDF
    doc = SimpleDocTemplate(output_pdf_path, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"<b>{trade} – Exercise {ex_no} ({year})</b>", styles["Title"]))
    elements.append(Paragraph(f"<b>Aim:</b> {aim}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Tolerance:</b> {tolerance}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    # Add job image
    image_path = os.path.join("static", "uploads", workorder.job_image or "")
    if os.path.exists(image_path):
        img = RLImage(image_path, width=300, height=200, kind='proportional')
        elements.append(img)
        elements.append(Spacer(1, 12))
    else:
        elements.append(Paragraph("⚠️ Job image not found.", styles["Normal"]))

    # Helper to render each table section
    def feature_table(title, labels, data, headers):
        elements.append(Paragraph(title, styles["Heading3"]))
        table_data = [headers] + [[labels[i], data[i][0], data[i][1]] for i in range(len(labels))]
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("BACKGROUND", (0,0), (-1,0), colors.lightblue),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 10))

    # Build feature tables
    feature_table("📐 Dimensional Features (A–J)", list("ABCDEFGHIJ"), list(zip(size_list, size_marks)), ["Label", "Size", "Marks"])
    feature_table("⚙️ Subjective Features (K–O)", list("KLMNO"), list(zip(subj_ops, subj_marks)), ["Label", "Feature", "Marks"])
    feature_table("🤝 Attitude Features (P–S)", list("PQRS"), list(zip(attitude_traits, attitude_marks)), ["Label", "Trait", "Marks"])

    doc.build(elements)