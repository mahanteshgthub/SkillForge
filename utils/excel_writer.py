
from openpyxl import load_workbook
from flask import current_app
import os  # ✅ This line is required and must be present
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage

#Populates a predifined Excel Template with dynamic data
def fill_excel_template(work, save_path, trainees=None, unit=None, dates=None):
    template_path = os.path.join(current_app.root_path, "static", "templates", "workorder_template.xlsx")
    print("Using template:", template_path)
    wb = load_workbook(template_path)
    ws = wb.active

    # ✅ Insert trainee names and grades first
    if trainees:
        print("Inserting trainees into Excel...")
        for i, trainee in enumerate(trainees[:20]):
            print(f"Row {42+i}: {trainee.name} - {trainee.grade}")
            ws[f"B{42+i}"] = trainee.name
            ws[f"AH{42+i}"] = trainee.grade

    # 🖼️ Job image
    MAX_WIDTH = 500
    MAX_HEIGHT = 475
    if work.job_image:
        image_path = os.path.join(current_app.root_path, "static", "uploads", work.job_image)
        if os.path.exists(image_path):
            with PILImage.open(image_path) as pil_img:
                orig_width, orig_height = pil_img.size
                scale_ratio = min(MAX_WIDTH / orig_width, MAX_HEIGHT / orig_height, 1)
                final_width = int(orig_width * scale_ratio)
                final_height = int(orig_height * scale_ratio)

            img = ExcelImage(image_path)
            img.width = final_width
            img.height = final_height
            ws.add_image(img, "B5")

    # 🧾 Header fields
    ws["C3"] = work.trade
    ws["E3"] = work.year
    print(f"Exercise No: '{work.exercise_no}'")
    ws["Q3"] = work.exercise_no
    ws["U3"] = work.aim
    ws["Z16"] = work.common_tolerance

    if unit:
        print(f"Inserting Trainer Unit info: ITI={unit.iti_name}, Shift={unit.shift_number}, Unit={unit.unit_number}")
        ws["Q1"] = unit.iti_name or ""
        ws["G3"] = unit.shift_number or ""
        ws["J3"] = unit.unit_number or ""
    else:
        print("No Trainer Unit found for this role.")
        ws["Q1"] = ""
        ws["G3"] = ""
        ws["J3"] = ""

    # Insert dates if present
    if dates:
        if "start" in dates:
            ws["AE37"] = dates["start"]
        if "end" in dates:
            ws["AE38"] = dates["end"]




    # 📏 Dimensional Features (A–J)
    dim_labels = "ABCDEFGHIJ"
    for i, label in enumerate(dim_labels):
        f = next((feat for feat in work.dimensional_features if feat.label == label), None)
        if f:
            ws[f"S{6+i}"] = f.size
            ws[f"AD{6+i}"] = f.marks

    # 🔧 Subjective Features (K–O)
    subj_labels = "KLMNO"
    for i, label in enumerate(subj_labels):
        f = next((feat for feat in work.subjective_features if feat.label == label), None)
        if f:
            ws[f"S{20+i}"] = f.operation
            ws[f"AB{20+i}"] = f.marks

    # 🧠 Attitude Features (P–S)
    att_labels = "PQRS"
    for i, label in enumerate(att_labels):
        f = next((feat for feat in work.attitude_features if feat.label == label), None)
        if f:
            ws[f"S{27+i}"] = f.trait
            ws[f"Z{27+i}"] = f.marks
    
    

    # ✅ Insert data into named ranges
    # ObjMarks → A41:J41
    obj_marks = [feat.marks for feat in work.dimensional_features]
    for i, val in enumerate(obj_marks):
        ws.cell(row=41, column=1+i).value = val  # columns A(1) to J(10)

    # SubMarks → K41:O41
    sub_marks = [feat.marks for feat in work.subjective_features]
    for i, val in enumerate(sub_marks):
        ws.cell(row=41, column=11+i).value = val  # columns K(11) to O(15)

    # AspectMarks → P41:S41
    att_marks = [feat.marks for feat in work.attitude_features]
    for i, val in enumerate(att_marks):
        ws.cell(row=41, column=16+i).value = val  # columns P(16) to S(19)
      
    wb.save(save_path)